import datetime
import math
import base64
from sqlite3 import IntegrityError
from flask import Blueprint, jsonify, request
from models import DailyAttendance, Employee, PayrollItem, PayrollSummary
from services.payroll_service import save_monthly_payroll, pph_21
from extensions import db
from io import BytesIO

from flask import send_file

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

payroll_bp = Blueprint(
    "payroll",
    __name__,
    url_prefix="/api"
)

@payroll_bp.route("/payroll/add-item", methods=["POST"])
def add_payroll_item():
    data = request.get_json() or {}

    emp_id = data.get("emp_id")
    periode = data.get("periode")
    item_type = data.get("type")
    name = data.get("name")
    kasbon = data.get("kasbon")
    cicilan = data.get("cicilan")
    rate = float(data.get("rate", 0))
    qty = float(data.get("qty", 1))
    amount = float(data.get("amount", 0))

    print("TEST: ",kasbon,"-",cicilan)

    # Cari employee
    emp = db.session.get(Employee, emp_id)

    if not emp:
        return jsonify({"error": "Karyawan tidak ditemukan"}), 404

    # Cari / buat summary
    summary = PayrollSummary.query.filter_by(
        employee_id=emp_id,
        periode=periode
    ).first()

    if not summary:
        summary = PayrollSummary(
            employee_id=emp_id,
            periode=periode,
            basic_salary=emp.basic_salary or 0,
            total_earning=emp.basic_salary or 0,
            total_deduction=0,
            thp=emp.basic_salary or 0
        )

        db.session.add(summary)
        db.session.flush()

    # Tambahkan item
    new_item = PayrollItem(
        payroll_summary_id=summary.id,
        name=name,
        type=item_type,
        rate=rate,
        qty=qty,
        unit="unit",
        amount=amount
    )

    db.session.add(new_item)
    db.session.flush()

    # Hitung ulang SEMUA komponen
    all_items = summary.items

    total_tunjangan = sum(
        item.amount for item in all_items
        if item.type == "tunjangan"
    )

    total_bonus = sum(
        item.amount for item in all_items
        if item.type == "bonus"
    )

    total_potongan = sum(
        item.amount for item in all_items
        if item.type == "potongan"
    )

    # Update summary
    summary.total_earning = (
        summary.basic_salary
        + total_tunjangan
        + total_bonus
    )

    summary.total_deduction = total_potongan

    summary.thp = (
        summary.total_earning
        - summary.total_deduction
    )

    summary.pph21 = pph_21(
        summary.thp,
        emp.gender,
        emp.married_status,
        emp.dependents
    )

    summary.thp_after_tax = (
        summary.thp - summary.pph21
    )

    summary.kasbon = kasbon
    summary.cicilan = cicilan

    db.session.commit()

    return jsonify({
        "success": True,
        "item_id": new_item.id,
        "total_tunjangan": total_tunjangan,
        "total_bonus": total_bonus,
        "total_potongan": total_potongan,
        "total_earning": summary.total_earning,
        "total_deduction": summary.total_deduction,
        "thp": summary.thp,
        "pph21": summary.pph21,
        "thp_after_tax": summary.thp_after_tax
    }), 200

@payroll_bp.route("/payroll/delete-item", methods=["POST"])
def delete_payroll_item():
    data = request.get_json() or {}

    item_id = data.get("item_id")

    item = db.session.get(PayrollItem, item_id)

    if not item:
        return jsonify({
            "error": "Item payroll tidak ditemukan"
        }), 404

    # Simpan summary sebelum item dihapus
    summary = item.summary

    # Hapus item
    db.session.delete(item)
    db.session.flush()

    # Ambil semua item yang tersisa
    all_items = summary.items

    total_tunjangan = sum(
        i.amount for i in all_items
        if i.type == "tunjangan"
    )

    total_bonus = sum(
        i.amount for i in all_items
        if i.type == "bonus"
    )

    total_potongan = sum(
        i.amount for i in all_items
        if i.type == "potongan"
    )

    # Hitung ulang summary
    summary.total_earning = (
        summary.basic_salary
        + total_tunjangan
        + total_bonus
    )

    summary.thp = (
        summary.total_earning
        - summary.total_deduction
    )

    summary.pph21 = pph_21(
        summary.thp,
        summary.employee.gender,
        summary.employee.married_status,
        summary.employee.dependents
    )

    summary.thp_after_tax = (
        summary.thp - summary.pph21
    )

    summary.total_deduction = total_potongan

    db.session.commit()

    return jsonify({
        "success": True,
        "item_id": item_id,
        "total_penambahan": total_tunjangan,
        "total_potongan": total_potongan,
        "thp": summary.thp,
        "pph21": summary.pph21,
        "thp_after_tax": summary.thp_after_tax
    }), 200

def add_payroll_item_internal(
    emp_id,
    periode,
    item_type,
    name,
    rate=0,
    qty=1,
    amount=0
):
    emp = db.session.get(Employee, emp_id)

    if not emp:
        raise ValueError("Karyawan tidak ditemukan")

    summary = PayrollSummary.query.filter_by(
        employee_id=emp_id,
        periode=periode
    ).first()

    if not summary:
        summary = PayrollSummary(
            employee_id=emp_id,
            periode=periode,
            basic_salary=emp.basic_salary or 0,
            total_earning=emp.basic_salary or 0,
            total_deduction=0,
            thp=emp.basic_salary or 0
        )

        db.session.add(summary)
        db.session.flush()

    new_item = PayrollItem(
        payroll_summary_id=summary.id,
        name=name,
        type=item_type,
        rate=float(rate or 0),
        qty=float(qty or 1),
        unit="unit",
        amount=float(amount or 0)
    )

    db.session.add(new_item)
    db.session.flush()

    return new_item

@payroll_bp.route("/payroll/generate", methods=["POST"])
def generate_payroll():
    try:
        data = request.get_json(silent=True) or {}

        periode = (data.get("periode") or "").strip()

        # ==========================================
        # VALIDASI PERIODE
        # ==========================================
        if not periode:
            return jsonify({
                "message": "Bulan/periode wajib dipilih"
            }), 400

        try:
            datetime.datetime.strptime(periode, "%Y-%m")
        except ValueError:
            return jsonify({
                "message": "Format periode harus YYYY-MM"
            }), 400


        # ==========================================
        # AMBIL SEMUA KARYAWAN
        # ==========================================
        employees = Employee.query.all()

        if not employees:
            return jsonify({
                "message": "Tidak ada data karyawan"
            }), 404


        created_count = 0
        skipped_count = 0


        # ==========================================
        # BUAT PAYROLL UNTUK SETIAP KARYAWAN
        # ==========================================
        for employee in employees:

            # ==========================================
            # CEK APAKAH PAYROLL SUDAH ADA
            # ==========================================
            existing_summary = PayrollSummary.query.filter_by(
                employee_id=employee.id,
                periode=periode
            ).first()

            print(employee.name, ": ", existing_summary)

            if existing_summary:
                skipped_count += 1
                continue


            # ==========================================
            # DATA BASIC SALARY
            # ==========================================
            basic_salary = float(employee.basic_salary or 0)

            print("GAPOK -> ", basic_salary)


            # ==========================================
            # BUAT PAYROLL SUMMARY DASAR
            # ==========================================
            new_summary = PayrollSummary(
                employee_id=employee.id,
                periode=periode,
                basic_salary=basic_salary,
                total_earning=basic_salary,
                total_deduction=0,
                thp=basic_salary,
                pph21=0,
                thp_after_tax=basic_salary,
                status_wa="Pending"
            )

            db.session.add(new_summary)
            db.session.flush()


            # ==========================================
            # AMBIL DATA ABSENSI
            # ==========================================
            attendance = DailyAttendance.query.filter(
                DailyAttendance.user_id == employee.user_id,
                DailyAttendance.date.like(f"{periode}-%")
            ).all()

            qty_absen = 0
            qty_terlambat = 0

            for record in attendance:

                if record.status == "A":
                    qty_absen += 1

                elif record.status == "T":
                    qty_terlambat += 1


            print("ABSEN -> ", qty_absen)
            print("TERLAMBAT -> ", qty_terlambat)


            # ==========================================
            # RATE
            # ==========================================
            rate_absen = math.floor(basic_salary / 30)
            rate_terlambat = 25000


            # ==========================================
            # POTONGAN ABSEN
            # ==========================================
            if qty_absen > 0 and rate_absen > 0:

                add_payroll_item_internal(
                    employee.id,
                    periode,
                    "potongan",
                    "Potongan Absen",
                    rate=rate_absen,
                    qty=qty_absen,
                    amount=rate_absen * qty_absen
                )

                print("Potongan absen sudah dihitung")


            # ==========================================
            # POTONGAN TERLAMBAT
            # ==========================================
            if qty_terlambat > 0 and rate_terlambat > 0:

                add_payroll_item_internal(
                    employee.id,
                    periode,
                    "potongan",
                    "Potongan Terlambat",
                    rate=rate_terlambat,
                    qty=qty_terlambat,
                    amount=rate_terlambat * qty_terlambat
                )

                print("Potongan terlambat sudah dihitung")


            # ==========================================
            # HITUNG ULANG DARI PAYROLL ITEM
            # ==========================================
            all_items = new_summary.items

            total_tunjangan = sum(
                item.amount
                for item in all_items
                if item.type == "tunjangan"
            )

            total_bonus = sum(
                item.amount
                for item in all_items
                if item.type == "bonus"
            )

            total_potongan = sum(
                item.amount
                for item in all_items
                if item.type == "potongan"
            )


            # ==========================================
            # TOTAL EARNING
            # ==========================================
            new_summary.total_earning = (
                basic_salary
                + total_tunjangan
                + total_bonus
            )


            # ==========================================
            # TOTAL DEDUCTION
            # ==========================================
            new_summary.total_deduction = total_potongan


            # ==========================================
            # THP
            # ==========================================
            new_summary.thp = (
                new_summary.total_earning
                - new_summary.total_deduction
            )

            print("TOTAL EARNING -> ", new_summary.total_earning)
            print("TOTAL DEDUCTION -> ", new_summary.total_deduction)
            print("THP -> ", new_summary.thp)


            # ==========================================
            # PPH 21
            # ==========================================
            new_summary.pph21 = pph_21(
                new_summary.thp,
                employee.gender,
                employee.married_status,
                employee.dependents
            )

            print("PPH21 -> ", new_summary.pph21)


            # ==========================================
            # THP AFTER TAX
            # ==========================================
            new_summary.thp_after_tax = (
                new_summary.thp
                - new_summary.pph21
            )

            print(
                "THP AFTER TAX -> ",
                new_summary.thp_after_tax
            )


            created_count += 1

        db.session.commit()

        return jsonify({
            "message": "Payroll berhasil dibuat",
            "created": created_count,
            "skipped": skipped_count
        }), 200

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "message": f"Gagal membuat payroll: {str(e)}"
        }), 500

@payroll_bp.route("/payroll/delete", methods=["DELETE"])
def delete_payroll():
    try:
        data = request.get_json(silent=True) or {}

        periode = (data.get("periode") or "").strip()

        # ==========================================
        # VALIDASI PERIODE
        # ==========================================
        if not periode:
            return jsonify({
                "message": "Bulan/periode wajib dipilih"
            }), 400

        try:
            datetime.datetime.strptime(periode, "%Y-%m")
        except ValueError:
            return jsonify({
                "message": "Format periode harus YYYY-MM"
            }), 400


        # ==========================================
        # CARI PAYROLL BERDASARKAN PERIODE
        # ==========================================
        payrolls = PayrollSummary.query.filter_by(
            periode=periode
        ).all()

        if not payrolls:
            return jsonify({
                "message": f"Tidak ada payroll untuk periode {periode}"
            }), 404


        jumlah_dihapus = len(payrolls)


        # ==========================================
        # HAPUS PAYROLL
        # ==========================================
        for payroll in payrolls:
            db.session.delete(payroll)


        db.session.commit()


        return jsonify({
            "message": f"Payroll periode {periode} berhasil dihapus",
            "periode": periode,
            "deleted": jumlah_dihapus
        }), 200


    except Exception as e:

        db.session.rollback()

        return jsonify({
            "message": f"Gagal menghapus payroll: {str(e)}"
        }), 500

@payroll_bp.route("/payroll/slip/<int:employee_id>")
def get_payroll_slip(employee_id):

    periode = request.args.get("periode")

    summaries = PayrollSummary.query.filter_by(
        periode=periode
    ).all()

    attendance_records = DailyAttendance.query.filter(
        DailyAttendance.date.like(f"{periode}-%")
    ).all()

    attendance_summary = {}

    for record in attendance_records:

        user_id = str(
            record.user_id
        ).strip()

        if user_id not in attendance_summary:

            attendance_summary[user_id] = {
                "H": 0,
                "T": 0,
                "S": 0,
                "A": 0,
                "L": 0
            }

        if record.status in attendance_summary[user_id]:

            attendance_summary[user_id][
                record.status
            ] += 1

    employees_data = []

    for s in summaries:

        emp = s.employee

        if not emp:
            continue

        # ======================================================
        # LIST PAYROLL ITEMS
        # ======================================================

        tunjangan_list = [
            {
                "id": i.id,
                "name": i.name,
                "amount": i.amount
            }
            for i in s.items
            if i.type == "tunjangan"
        ]

        bonus_list = [
            {
                "id": i.id,
                "name": i.name,
                "amount": i.amount
            }
            for i in s.items
            if i.type == "bonus"
        ]

        potongan_list = [
            {
                "id": i.id,
                "name": i.name,
                "amount": i.amount
            }
            for i in s.items
            if i.type == "potongan"
        ]

        # ======================================================
        # AMBIL REKAP ABSENSI EMPLOYEE
        # ======================================================

        emp_user_id = str(
            emp.user_id
        ).strip()

        att = attendance_summary.get(
            emp_user_id,
            {
                "H": 0,
                "T": 0,
                "S": 0,
                "A": 0,
                "L": 0
            }
        )

        # ======================================================
        # HITUNG PPH 21
        # ======================================================

        # pph21 = pph_21(
        #     s.thp,
        #     emp.married_status,
        #     emp.dependents
        # )

        # thp_after_tax = s.thp - pph21

        # ======================================================
        # DATA EMPLOYEE
        # ======================================================

        employees_data.append({
            "id": str(emp.id),
            "user_id": emp_user_id,
            "name": emp.name or "-",
            "position": emp.position or "-",
            "basic_salary": s.basic_salary,
            "thp": s.thp,
            "pph21": s.pph21,
            "kasbon": s.kasbon,
            "cicilan": s.cicilan,
            "thp_after_tax": s.thp_after_tax,
            "tunjanganList": tunjangan_list,
            "bonusList": bonus_list,
            "potonganList": potongan_list,
            "att": {
                "H": att["H"],
                "T": att["T"],
                "S": att["S"],
                "A": att["A"],
                "L": att["L"]
            }
        })

    
    employees_data.sort(
        key=lambda x: (
            (x["position"] or "-").lower(),
            (x["name"] or "-").lower()
        )
    )

    # ==========================================================
    # 6. RENDER
    # ==========================================================

    return jsonify(
        employees_json=employees_data,
    )

@payroll_bp.route("/payroll/slip/all")
def download_all_payroll():

    periode = request.args.get("periode")

    if not periode:
        return jsonify({
            "error": "Parameter periode wajib diisi."
        }), 400

    # ==========================================================
    # KONFIGURASI SURAT
    # ==========================================================

    COMPANY_CITY = "Cibubur"
    BANK_DESTINATION = "PT Bank UOB Indonesia Cab. Cibubur"
    DIRECTOR_NAME = "Bonatua Silalahi"
    COMPANY_ACCOUNT_NAME = "PT. HAMASA IPARNA MANDIRI"
    COMPANY_ACCOUNT_NUMBER = "5403003310"

    # ==========================================================
    # 1. VALIDASI PERIODE
    # ==========================================================

    try:
        year, month = map(int, periode.split("-"))
        if month < 1 or month > 12:
            raise ValueError

    except ValueError:

        return jsonify({
            "error": "Format periode harus YYYY-MM."
        }), 400

    # ==========================================================
    # 2. NAMA BULAN
    # ==========================================================

    nama_bulan = [ "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

    nama_bulan_periode = nama_bulan[month - 1]

    # ==========================================================
    # 3. TANGGAL SURAT
    # ==========================================================

    today = datetime.date.today()

    tanggal_surat = (
        f"{COMPANY_CITY}, "
        f"{today.day} "
        f"{nama_bulan[today.month - 1]} "
        f"{today.year}"
    )

    # ==========================================================
    # 4. AMBIL DATA PAYROLL
    # ==========================================================

    summaries = PayrollSummary.query.filter_by(
        periode=periode
    ).all()

    if not summaries:
        return jsonify({
            "error": (
                f"Tidak ada data payroll "
                f"untuk periode {periode}."
            )
        }), 404

    # ==========================================================
    # 5. SIAPKAN DATA EMPLOYEE
    # ==========================================================

    employees_data = []
    for s in summaries:
        emp = s.employee
        if not emp:
            continue

        account_number = emp.no_rekening or "-"
        bank_name = emp.nama_bank or "-"

        employees_data.append({
            "name": emp.name or "-",
            "account_number": account_number,
            "bank_name": bank_name,
            "currency": "IDR",
            "amount": s.thp_after_tax or 0
        })

    # ==========================================================
    # 6. SORT EMPLOYEE BERDASARKAN NAMA
    # ==========================================================

    employees_data.sort(
        key=lambda x: (
            x["name"] or "-"
        ).lower()
    )

    # ==========================================================
    # 7. TOTAL PAYROLL
    # ==========================================================

    total_payroll = sum(
        emp["amount"]
        for emp in employees_data
    )

    # ==========================================================
    # 8. FUNGSI TERBILANG
    # ==========================================================

    def terbilang(n):
        angka = [ "", "Satu", "Dua", "Tiga", "Empat", "Lima", "Enam", "Tujuh", "Delapan", "Sembilan", "Sepuluh", "Sebelas" ]

        n = int(n)

        if n < 12:
            return angka[n]

        elif n < 20:
            return ( terbilang(n - 10) + " Belas" )

        elif n < 100:
            return ( terbilang(n // 10) + " Puluh " + terbilang(n % 10) )

        elif n < 200:
            return ( "Seratus " + terbilang(n - 100) )

        elif n < 1000:
            return ( terbilang(n // 100) + " Ratus " + terbilang(n % 100) )

        elif n < 2000:
            return ( "Seribu " + terbilang(n - 1000) )

        elif n < 1_000_000:
            return ( terbilang(n // 1000) + " Ribu " + terbilang(n % 1000) )

        elif n < 1_000_000_000:
            return ( terbilang(n // 1_000_000) + " Juta " + terbilang(n % 1_000_000) )

        elif n < 1_000_000_000_000:
            return ( terbilang(n // 1_000_000_000) + " Miliar " + terbilang(n % 1_000_000_000) )

        elif n < 1_000_000_000_000_000:
            return ( terbilang(n // 1_000_000_000_000) + " Triliun " + terbilang(n % 1_000_000_000_000) )

        return "Jumlah terlalu besar"

    # ----------------------------------------------------------
    # Bersihkan spasi berlebih
    # ----------------------------------------------------------

    total_terbilang = " ".join(
        terbilang(total_payroll).split()
    )

    total_terbilang = (
        total_terbilang
        + " Rupiah"
    )

    # ==========================================================
    # 9. BUAT WORKBOOK
    # ==========================================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Payroll"

    # ==========================================================
    # 10. STYLE
    # ==========================================================

    title_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    total_fill = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9"
    )

    title_font = Font(
        color="FFFFFF",
        bold=True,
        size=16
    )

    header_font = Font(
        color="000000",
        bold=True,
        size=12
    )

    bold_font = Font(
        bold=True
    )

    normal_font = Font(
        size=12
    )

    thin_side = Side(
        style="medium",
        color="000000"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    right = Alignment(
        horizontal="right",
        vertical="center",
        wrap_text=True
    )

    # ==========================================================
    # 11. TANGGAL SURAT
    # ==========================================================

    ws.merge_cells("A1:F1")

    ws["A1"] = tanggal_surat
    ws["A1"].font = Font(size=12)
    ws["A1"].alignment = left

    # ==========================================================
    # 12. KEPADA YTH.
    # ==========================================================

    ws.merge_cells("A3:F3")

    ws["A3"] = (
        f"Kepada Yth.\n"
        f"{BANK_DESTINATION}\n\n"
        f"Saya {DIRECTOR_NAME}, sebagai Direktur dari:\n\n"
        f"Nama Rekening   : {COMPANY_ACCOUNT_NAME}\n"
        f"Nomor Rekening  : {COMPANY_ACCOUNT_NUMBER}\n\n"
        f"Memberikan instruksi pindah buku untuk pembayaran "
        f"Payroll ke rekening {BANK_DESTINATION} kepada nasabah "
        f"sesuai dengan list di bawah ini:"
    )

    ws["A3"].font = normal_font
    ws["A3"].alignment = left

    # ==========================================================
    # 13. TINGGI ROW SURAT
    # ==========================================================

    ws.row_dimensions[3].height = 160

    # ==========================================================
    # 14. JUDUL PAYROLL
    # ==========================================================

    ws.merge_cells("A5:F5")

    ws["A5"] = (
        f"Gaji Bulan "
        f"{nama_bulan_periode} "
        f"{year}"
    )

    ws["A5"].font = Font(size=12)

    ws["A5"].alignment = center

    ws.row_dimensions[5].height = 32

    # ==========================================================
    # 15. HEADER TABEL
    # ==========================================================

    headers = [
        "No",
        "Nama",
        "No. Rekening",
        "Nama Bank",
        "Mata Uang",
        "Nominal"
    ]

    header_row = 6

    for col, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=header_row,
            column=col,
            value=header
        )

        # Header dibuat seperti baris tabel biasa
        cell.font = Font(
            bold=True,
            size=11
        )

        cell.alignment = center
        cell.border = border

    ws.row_dimensions[header_row].height = 30

    # ==========================================================
    # 16. DATA KARYAWAN
    # ==========================================================

    start_data_row = 7

    for index, emp in enumerate(
        employees_data,
        start=1
    ):

        row = start_data_row + index - 1

        values = [
            index,
            emp["name"],
            emp["account_number"],
            emp["bank_name"],
            "IDR",
            emp["amount"]
        ]

        for col, value in enumerate(
            values,
            start=1
        ):

            cell = ws.cell(
                row=row,
                column=col,
                value=value
            )

            cell.border = border
            cell.font = Font(size=12)

            if col in [1, 5]:
                cell.alignment = center

            elif col == 6:
                cell.alignment = right

            else:
                cell.alignment = left

        ws.row_dimensions[row].height = 18

    # ==========================================================
    # 17. FORMAT NOMINAL
    # ==========================================================

    currency_format = '#,##0'

    for row in range(
        start_data_row,
        start_data_row + len(employees_data)
    ):

        ws.cell(
            row=row,
            column=6
        ).number_format = currency_format

    # ==========================================================
    # 18. TOTAL
    # ==========================================================

    total_row = (
        start_data_row
        + len(employees_data)
    )

    # Merge A:E
    ws.merge_cells(
        start_row=total_row,
        start_column=1,
        end_row=total_row,
        end_column=5
    )

    total_label = ws.cell(
        row=total_row,
        column=1,
        value="TOTAL"
    )

    # total_label.fill = total_fill
    total_label.font = bold_font
    total_label.alignment = center
    total_label.border = border

    # Nominal total
    total_cell = ws.cell(
        row=total_row,
        column=6,
        value=total_payroll
    )

    # total_cell.fill = total_fill
    total_cell.font = bold_font
    total_cell.alignment = right
    total_cell.border = border
    total_cell.number_format = currency_format

    # ==========================================================
    # 19. BORDER TOTAL
    # ==========================================================

    for col in range(1, 7):

        cell = ws.cell(
            row=total_row,
            column=col
        )

        cell.border = border

        # if col <= 5:
        #     cell.fill = total_fill

    ws.row_dimensions[total_row].height = 28

    # ==========================================================
    # 20. TOTAL ANGKA
    # ==========================================================

    amount_number_row = total_row + 2

    ws.merge_cells(
        start_row=amount_number_row,
        start_column=1,
        end_row=amount_number_row,
        end_column=6
    )

    ws.cell(
        row=amount_number_row,
        column=1,
        value=(
            f"Total Transaksi : Rp{total_payroll:,.0f}"
        )
    )

    ws.cell(
        row=amount_number_row,
        column=1
    ).font = Font(
        bold=True,
        size=12
    )

    ws.cell(
        row=amount_number_row,
        column=1
    ).alignment = left

    # ==========================================================
    # 21. TOTAL TERBILANG
    # ==========================================================

    amount_words_row = total_row + 3

    ws.merge_cells(
        start_row=amount_words_row,
        start_column=1,
        end_row=amount_words_row,
        end_column=6
    )

    ws.cell(
        row=amount_words_row,
        column=1,
        value=total_terbilang
    )

    ws.cell(
        row=amount_words_row,
        column=1
    ).font = Font(
        italic=True,
        bold=True,
        size=12
    )

    ws.cell(
        row=amount_words_row,
        column=1
    ).alignment = left

    # ==========================================================
    # 22. PENUTUP
    # ==========================================================

    closing_row = total_row + 5

    ws.merge_cells(
        start_row=closing_row,
        start_column=1,
        end_row=closing_row,
        end_column=6
    )

    closing_cell = ws.cell(
        row=closing_row,
        column=1,
        value=(
            "Mohon agar transaksi ini dapat dijalankan "
            "segera. Terima kasih atas bantuannya."
        )
    )

    closing_cell.font = Font(size=12)
    closing_cell.alignment = left


    # ==========================================================
    # 23. HORMAT SAYA
    # ==========================================================

    signature_row = closing_row + 3

    ws.merge_cells(
        start_row=signature_row,
        start_column=1,
        end_row=signature_row,
        end_column=6
    )

    signature_cell = ws.cell(
        row=signature_row,
        column=1,
        value="Hormat Saya,"
    )

    signature_cell.font = Font(size=12)
    signature_cell.alignment = left


    # ==========================================================
    # 24. NAMA DIREKTUR
    # ==========================================================

    director_row = signature_row + 5

    ws.merge_cells(
        start_row=director_row,
        start_column=1,
        end_row=director_row,
        end_column=6
    )

    director_cell = ws.cell(
        row=director_row,
        column=1,
        value=DIRECTOR_NAME
    )

    director_cell.font = Font(
        bold=True,
        size=12
    )

    director_cell.alignment = left


    # ==========================================================
    # 25. JABATAN
    # ==========================================================

    position_row = director_row + 1

    ws.merge_cells(
        start_row=position_row,
        start_column=1,
        end_row=position_row,
        end_column=6
    )

    position_cell = ws.cell(
        row=position_row,
        column=1,
        value="Direktur"
    )

    position_cell.font = Font(
        size=12
    )

    position_cell.alignment = left


    # ==========================================================
    # 26. COLUMN WIDTH
    # ==========================================================

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12


    # ==========================================================
    # 27. GRIDLINES
    # ==========================================================

    ws.sheet_view.showGridLines = False


    # ==========================================================
    # 28. FREEZE HEADER
    # ==========================================================

    # Tidak menggunakan freeze/floating header


    # ==========================================================
    # 29. PRINT SETTING
    # ==========================================================

    ws.page_setup.orientation = "landscape"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True


    # ==========================================================
    # MARGIN
    # ==========================================================

    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2


    # ==========================================================
    # PRINT AREA
    # ==========================================================

    ws.print_area = (
        f"A1:F{position_row}"
    )


    # ==========================================================
    # 30. PAGE FOOTER
    # ==========================================================

    ws.oddFooter.center.text = (
        "Payroll - "
        f"{nama_bulan_periode} {year}"
    )


    # ==========================================================
    # 31. DOWNLOAD
    # ==========================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    filename = f"Payroll_{periode}.xlsx"

    file_data = base64.b64encode(
        output.read()
    ).decode("utf-8")

    return jsonify({
        "success": True,
        "filename": filename,
        "data": file_data
    })